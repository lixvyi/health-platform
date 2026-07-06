# -*- coding: utf-8 -*-
"""Parse 年度数据 (*.xlsx) from NBS into consolidated JSON."""
import json
import os
import re
import shutil
import sys

from openpyxl import load_workbook

FILE_TITLES = {
    1: "卫生人员",
    2: "每千人口卫生资源",
    3: "医疗卫生机构数",
    4: "医疗卫生机构床位",
    5: "城乡床位分布",
    6: "平均住院日",
    7: "病床使用率",
    8: "医院平均住院日",
    9: "医院服务效率",
    10: "基层卫生服务",
    11: "甲乙类传染病发病数",
    12: "甲乙类传染病死亡数",
    13: "甲乙类传染病发病率",
    14: "甲乙类传染病死亡率",
    15: "城市居民疾病死亡率",
    16: "城市居民死因构成",
    17: "城市居民死因顺位",
    18: "农村居民疾病死亡率",
    19: "农村居民死因构成",
    20: "农村居民死因顺位",
    21: "孕产妇和婴儿死亡率",
    22: "儿童免疫接种率",
    23: "孕产妇保健率",
    24: "行政区划",
    25: "卫生总费用",
}


def parse_year_header(cell):
    if not cell:
        return None
    s = str(cell).strip()
    m = re.match(r"(\d{4})", s)
    return m.group(1) if m else None


def parse_file(path: str, index: int) -> dict | None:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 4:
        return None

    meta_db = str(rows[0][0] or "")
    meta_time = str(rows[1][0] or "")
    headers = rows[2]
    years = [parse_year_header(h) for h in headers[1:] if h]

    indicators = []
    for row in rows[3:]:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        if "国家统计局" in name or name.startswith("注"):
            continue
        values = {}
        for yi, y in enumerate(years):
            if not y:
                continue
            val = row[yi + 1] if yi + 1 < len(row) else None
            if val is None or val == "":
                continue
            try:
                values[y] = float(val)
            except (TypeError, ValueError):
                s = str(val).strip()
                if s:
                    try:
                        values[y] = float(s)
                    except ValueError:
                        pass
        if values:
            indicators.append({"name": name, "values": values})

    if not indicators:
        return None

    return {
        "id": f"nbs-{index:02d}",
        "fileIndex": index,
        "title": FILE_TITLES.get(index, f"年度数据{index}"),
        "source": "国家统计局",
        "sourceUrl": "https://data.stats.gov.cn",
        "database": meta_db,
        "timeRange": meta_time,
        "indicators": indicators,
    }


def main():
    src_dir = sys.argv[1] if len(sys.argv) > 1 else "d:/"
    out_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "open-data", "stats")
    os.makedirs(out_dir, exist_ok=True)

    datasets = []
    for i in range(1, 26):
        src = os.path.join(src_dir, f"年度数据 ({i}).xlsx")
        if not os.path.exists(src):
            print("SKIP missing:", src)
            continue
        dst = os.path.join(out_dir, f"年度数据_{i:02d}.xlsx")
        shutil.copy2(src, dst)
        parsed = parse_file(src, i)
        if parsed:
            datasets.append(parsed)
            print("OK", i, parsed["title"], len(parsed["indicators"]))

    payload = {
        "source": "国家统计局",
        "sourceUrl": "https://data.stats.gov.cn",
        "attribution": "来源：国家统计局",
        "datasets": datasets,
    }
    json_path = os.path.join(out_dir, "nbs-health-stats.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    backend_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        "health-portal-backend",
        "src",
        "main",
        "resources",
        "data",
        "nbs-health-stats.json",
    )
    os.makedirs(os.path.dirname(backend_path), exist_ok=True)
    shutil.copy2(json_path, backend_path)
    print("WROTE", json_path)
    print("WROTE", backend_path)
    print("TOTAL datasets:", len(datasets))


if __name__ == "__main__":
    main()
