#!/usr/bin/env python3
"""真实健康资源预检、标准化与幂等导入工具。

默认仅执行 dry-run，不连接数据库、不修改原始文件。只有显式传入
``--apply`` 时才会连接数据库并写入迁移后的目标表。
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sys
import uuid
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Iterable

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
EXTERNAL_ROOT = ROOT / "data" / "external-import"
DEFAULT_OUTPUT = ROOT / "data" / "processed" / "health-resources"
POLICY_ROOT = ROOT / "data" / "policies"
POLICY_DRUG_CATALOG_CODE = "POLICY_NHSA_DRUG_CATALOG_2026"
POLICY_DRUG_CATALOG_NAME = "国家医保药品目录数据信息"

DATASET_META = {
    "hospitals": ("HOSPITAL_DIRECTORY_2024", "全国医院数据库"),
    "tertiary": ("PUBLIC_TERTIARY_HOSPITALS", "全国三级公立综合医院等级名单"),
    "grades": ("FUDAN_HOSPITAL_GRADES", "复旦医院等级分档"),
    "drugs": ("NATIONAL_DRUG_CATALOG_2025", "2025年国家医保药品目录"),
    "vaccine-schedule": ("VACCINE_SCHEDULE_2021", "国家免疫规划儿童免疫程序"),
    "vaccine-hiv": ("VACCINE_HIV_GUIDANCE_2021", "HIV感染母亲所生儿童接种建议"),
}


@dataclass
class DatasetResult:
    code: str
    name: str
    source_file: Path
    records: list[dict[str, Any]] = field(default_factory=list)
    errors: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    duplicate_count: int = 0

    def summary(self) -> dict[str, Any]:
        return {
            "datasetCode": self.code,
            "datasetName": self.name,
            "sourceFile": str(self.source_file.relative_to(ROOT)),
            "sourceSha256": sha256_file(self.source_file),
            "recordCount": len(self.records),
            "duplicateCount": self.duplicate_count,
            "errorCount": len(self.errors),
            "warnings": self.warnings,
        }


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def int_or_none(value: Any) -> int | None:
    text = clean(value).replace(",", "")
    if not text or not re.fullmatch(r"\d+(?:\.0+)?", text):
        return None
    return int(float(text))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def single_file(folder: str, pattern: str) -> Path:
    files = sorted((EXTERNAL_ROOT / folder).glob(pattern))
    if len(files) != 1:
        raise RuntimeError(f"{folder}/{pattern} 应有且仅有一个文件，实际为 {len(files)}")
    return files[0]


def policy_drug_catalog_file() -> Path | None:
    files = sorted(POLICY_ROOT.glob("*药品目录数据信息202606291435.csv"))
    return files[0] if files else None


def workbook(path: Path):
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def audit_dataset_summary(code: str) -> dict[str, Any]:
    report = DEFAULT_OUTPUT / "audit-report.json"
    if not report.exists():
        return {}
    try:
        payload = json.loads(report.read_text(encoding="utf-8"))
    except Exception:
        return {}
    for row in payload.get("datasets", []):
        if row.get("datasetCode") == code:
            return row
    return {}


def load_processed_dataset(key: str) -> DatasetResult | None:
    code, name = DATASET_META[key]
    path = DEFAULT_OUTPUT / f"{code}.json"
    if not path.exists():
        return None
    records = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(records, list):
        raise RuntimeError(f"{path.relative_to(ROOT)} 不是记录数组")

    error_path = DEFAULT_OUTPUT / f"{code}.errors.json"
    errors = []
    if error_path.exists():
        errors = json.loads(error_path.read_text(encoding="utf-8"))
        if not isinstance(errors, list):
            errors = []

    summary = audit_dataset_summary(code)
    result = DatasetResult(code, name, path)
    result.records = records
    result.errors = errors
    if key == "drugs":
        result.errors = []
        result.duplicate_count = count_duplicates(records, ("drugNumber", "drugName"))
        missing_dosage = sum(1 for record in records if not clean(record.get("dosageForm")))
        result.warnings = [f"{missing_dosage} 条记录没有独立剂型；保持空值，不从药名推断。"]
        merge_policy_drug_catalog(result)
    else:
        result.warnings = list(summary.get("warnings") or [])
        result.duplicate_count = int(summary.get("duplicateCount") or 0)
    return result


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            with path.open(encoding=encoding, newline="") as stream:
                return list(csv.DictReader(stream))
        except UnicodeDecodeError as exc:
            last_error = exc
    raise RuntimeError(f"无法读取 CSV 编码：{path.relative_to(ROOT)}") from last_error


def normalize_code(value: Any) -> str:
    text = clean(value)
    if re.fullmatch(r"\d+\.0+", text):
        return str(int(float(text)))
    return text


def parse_policy_drug_catalog() -> DatasetResult | None:
    path = policy_drug_catalog_file()
    if path is None:
        return None

    result = DatasetResult(POLICY_DRUG_CATALOG_CODE, POLICY_DRUG_CATALOG_NAME, path)
    source_rel = str(path.relative_to(ROOT))
    seen: Counter[tuple[str, str]] = Counter()
    rows = read_csv_rows(path)

    for index, row in enumerate(rows, 2):
        category_name = clean(row.get("结算项目名称")) or "医保药品"
        dosage_form = clean(row.get("剂型名称"))
        specification = clean(row.get("规格"))
        self_pay_ratio = clean(row.get("首先自负比例"))
        number = normalize_code(row.get("医疗项目编码"))
        insurance_type = clean(row.get("目录等级"))
        name = clean(row.get("药品名称"))
        manufacturer = clean(row.get("生产厂家"))

        if not name or not number:
            result.errors.append({
                "sourceFile": source_rel,
                "sourceSheet": "CSV",
                "sourceRow": index,
                "reason": "药品名称或医疗项目编码为空",
            })
            continue

        remark_parts = []
        if specification:
            remark_parts.append(f"规格：{specification}")
        if self_pay_ratio:
            remark_parts.append(f"首先自负比例：{self_pay_ratio}")
        if manufacturer:
            remark_parts.append(f"生产厂家：{manufacturer}")

        seen[(number, name)] += 1
        result.records.append({
            "categoryCode": category_name,
            "categoryName": category_name,
            "drugNumber": number,
            "drugName": name,
            "dosageForm": dosage_form,
            "insuranceType": insurance_type,
            "remark": "；".join(remark_parts),
            "catalogYear": 2026,
            "sourceFile": source_rel,
            "sourceSheet": "CSV",
            "sourceRow": index,
        })

    result.duplicate_count = sum(v - 1 for v in seen.values() if v > 1)
    result.warnings.append(
        f"追加 {len(result.records)} 条来自 data/policies 的医保药品目录数据信息；保留 CSV 来源行号。"
    )
    return result


def load_policy_drug_catalog() -> DatasetResult | None:
    csv_result = parse_policy_drug_catalog()
    path = DEFAULT_OUTPUT / f"{POLICY_DRUG_CATALOG_CODE}.json"
    if not path.exists():
        return csv_result

    records = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(records, list):
        raise RuntimeError(f"{path.relative_to(ROOT)} 不是记录数组")
    result = DatasetResult(POLICY_DRUG_CATALOG_CODE, POLICY_DRUG_CATALOG_NAME, path)
    result.records = records

    error_path = DEFAULT_OUTPUT / f"{POLICY_DRUG_CATALOG_CODE}.errors.json"
    if error_path.exists():
        errors = json.loads(error_path.read_text(encoding="utf-8"))
        result.errors = errors if isinstance(errors, list) else []
    if csv_result is not None:
        result.errors = csv_result.errors
    result.duplicate_count = count_duplicates(records, ("drugNumber", "drugName"))
    if csv_result is not None:
        result.warnings = csv_result.warnings
    else:
        result.warnings.append(
            f"追加 {len(result.records)} 条来自 data/policies 的医保药品目录数据信息；保留 CSV 来源行号。"
        )
    return result


def count_duplicates(records: list[dict[str, Any]], fields: tuple[str, ...]) -> int:
    seen = Counter(tuple(clean(record.get(field)) for field in fields) for record in records)
    return sum(count - 1 for count in seen.values() if count > 1)


def merge_policy_drug_catalog(result: DatasetResult) -> None:
    policy_result = load_policy_drug_catalog()
    if policy_result is None:
        result.warnings.append("未找到 data/policies 中的医保药品目录数据信息 CSV。")
        return
    result.records.extend(policy_result.records)
    result.errors.extend(policy_result.errors)
    result.duplicate_count += policy_result.duplicate_count
    result.warnings.extend(policy_result.warnings)


def parse_hospitals() -> DatasetResult:
    path = single_file("hospitals", "*.xlsx")
    result = DatasetResult("HOSPITAL_DIRECTORY_2024", "全国医院数据库", path)
    wb = workbook(path)
    ws = wb["Sheet1"]
    headers = [clean(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    expected = [
        "序号", "省", "市", "区县", "医院名称", "医院别名", "医院等级", "医院类型",
        "建院年份", "院长姓名", "经营方式", "是否医保", "床位数", "年门诊量",
        "医护人数", "医院科室", "电话", "邮箱", "医院地址", "邮编", "医院简介",
    ]
    if headers[: len(expected)] != expected:
        raise RuntimeError(f"医院表头不符合预期：{headers}")

    names: Counter[str] = Counter()
    source_rel = str(path.relative_to(ROOT))
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(clean(v) for v in row):
            continue
        name = clean(row[4])
        if not name:
            result.errors.append({"sourceRow": row_number, "reason": "医院名称为空"})
            continue
        names[name] += 1
        result.records.append({
            "sourceRecordNo": clean(row[0]),
            "province": clean(row[1]),
            "city": clean(row[2]),
            "district": clean(row[3]),
            "name": name,
            "aliasName": clean(row[5]),
            "level": clean(row[6]),
            "type": clean(row[7]),
            "foundedYear": clean(row[8]),
            "operationMode": clean(row[10]),
            "isInsurance": 1 if clean(row[11]) == "医保" else 0,
            "bedCount": int_or_none(row[12]),
            "annualVisits": int_or_none(row[13]),
            "medicalStaffCount": int_or_none(row[14]),
            "departments": clean(row[15]),
            "phone": clean(row[16]),
            "email": clean(row[17]),
            "address": clean(row[18]),
            "postalCode": clean(row[19]),
            "introduction": clean(row[20]),
            "sourceName": "用户提供的2024年全国医院数据库",
            "sourceFile": source_rel,
            "sourceSheet": ws.title,
            "sourceRow": row_number,
        })
    wb.close()
    result.duplicate_count = sum(count - 1 for count in names.values() if count > 1)
    if result.duplicate_count:
        result.warnings.append(
            f"存在 {result.duplicate_count} 条医院名称重复记录；按来源行保留，不能仅凭名称合并院区。"
        )
    return result


def parse_tertiary_hospitals() -> DatasetResult:
    path = single_file("insurance", "*.xlsx")
    result = DatasetResult("PUBLIC_TERTIARY_HOSPITALS", "全国三级公立综合医院等级名单", path)
    wb = workbook(path)
    if "医院明细" not in wb.sheetnames:
        raise RuntimeError("三级医院工作簿缺少“医院明细”工作表")
    ws = wb["医院明细"]
    source_rel = str(path.relative_to(ROOT))
    seen: Counter[tuple[str, str, str]] = Counter()
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        grade, province, hospital = map(clean, row[:3])
        if not any((grade, province, hospital)):
            continue
        if not all((grade, province, hospital)):
            result.errors.append({"sourceRow": row_number, "reason": "等级、省份或医院名称为空"})
            continue
        seen[(grade, province, hospital)] += 1
        result.records.append({
            "grade": grade,
            "province": province,
            "hospitalName": hospital,
            "sourceName": "用户提供的全国三级公立综合医院等级名单",
            "sourceFile": source_rel,
            "sourceSheet": ws.title,
            "sourceRow": row_number,
        })
    wb.close()
    result.duplicate_count = sum(v - 1 for v in seen.values() if v > 1)
    return result


def parse_hospital_grades() -> DatasetResult:
    path = single_file("rankings", "*.xlsx")
    result = DatasetResult("FUDAN_HOSPITAL_GRADES", "复旦医院等级分档", path)
    wb = workbook(path)
    ws = wb.active
    source_rel = str(path.relative_to(ROOT))
    seen: Counter[str] = Counter()
    allowed = {"A++++", "A+++", "A++", "A+", "A"}
    for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
        grade, hospital = map(clean, row[:2])
        if not any((grade, hospital)):
            continue
        if grade not in allowed or not hospital:
            result.errors.append({"sourceRow": row_number, "reason": "等级或医院名称无效"})
            continue
        seen[hospital] += 1
        result.records.append({
            "grade": grade,
            "hospitalName": hospital,
            "sourceName": "用户提供的复旦医院等级分档表",
            "sourceFile": source_rel,
            "sourceSheet": ws.title,
            "sourceRow": row_number,
        })
    wb.close()
    result.duplicate_count = sum(v - 1 for v in seen.values() if v > 1)
    result.warnings.append("该文件只有等级和医院名称，不含数字名次、专科或年份，禁止当作专科Top 10。")
    return result


def parse_drugs() -> DatasetResult:
    path = single_file("other", "1.*.xlsx")
    result = DatasetResult("NATIONAL_DRUG_CATALOG_2025", "2025年国家医保药品目录", path)
    wb = workbook(path)
    ws = wb.active
    source_rel = str(path.relative_to(ROOT))
    category_code = ""
    category_name = ""
    seen: Counter[tuple[str, str]] = Counter()
    missing_dosage = 0

    for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
        # 西药部分：A分类代码、C分类、L医保类别、N编号、R药名、Y剂型。
        if row_number < 2438:
            if clean(row[0]):
                category_code = clean(row[0])
            if clean(row[2]):
                category_name = clean(row[2])
            number = clean(row[13])
            name = clean(row[17])
            insurance_type = clean(row[11])
            dosage_form = clean(row[24])
            remark = ""
        else:
            # 中成药部分：A分类代码、D分类、R医保类别、T编号、V药名、AA备注。
            if clean(row[0]):
                category_code = clean(row[0])
            if clean(row[3]):
                category_name = clean(row[3])
            number = clean(row[19])
            name = clean(row[21])
            insurance_type = clean(row[17])
            dosage_form = ""  # 原表没有独立剂型列，禁止从名称猜测。
            remark = clean(row[26])

        if number == "编号" or name == "药品名称" or not number or not name:
            continue
        if not category_code or not category_name:
            result.errors.append({"sourceRow": row_number, "reason": "药品分类代码或分类为空"})
            continue
        if not dosage_form:
            missing_dosage += 1
        seen[(number, name)] += 1
        result.records.append({
            "categoryCode": category_code,
            "categoryName": category_name,
            "drugNumber": number,
            "drugName": name,
            "dosageForm": dosage_form,
            "insuranceType": insurance_type,
            "remark": remark,
            "catalogYear": 2025,
            "sourceFile": source_rel,
            "sourceSheet": ws.title,
            "sourceRow": row_number,
        })
    wb.close()
    result.duplicate_count = sum(v - 1 for v in seen.values() if v > 1)
    result.warnings.append(f"{missing_dosage} 条记录没有独立剂型；保持空值，不从药名推断。")
    return result


def vaccine_files() -> tuple[Path, Path]:
    files = sorted((EXTERNAL_ROOT / "other").glob("*疫苗*.xlsx"))
    if len(files) != 2:
        raise RuntimeError(f"疫苗 Excel 应为两个，实际为 {len(files)}")
    hiv = next((p for p in files if p.stem.endswith("(1)")), None)
    schedule = next((p for p in files if p != hiv), None)
    if not hiv or not schedule:
        raise RuntimeError("无法区分免疫程序表和 HIV 儿童接种建议表")
    return schedule, hiv


def parse_vaccine_schedule() -> DatasetResult:
    path, _ = vaccine_files()
    result = DatasetResult("VACCINE_SCHEDULE_2021", "国家免疫规划儿童免疫程序", path)
    wb = workbook(path)
    ws = wb.active
    age_headers = [clean(v) for v in next(ws.iter_rows(min_row=3, max_row=3, values_only=True))[5:20]]
    source_rel = str(path.relative_to(ROOT))
    for row_number, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
        vaccine_name = clean(row[1])
        if not vaccine_name:
            continue
        schedule = {
            age: clean(value)
            for age, value in zip(age_headers, row[5:20])
            if age and clean(value)
        }
        result.records.append({
            "preventableDisease": clean(row[0]),
            "vaccineName": vaccine_name,
            "route": clean(row[2]),
            "dose": clean(row[3]),
            "abbreviation": clean(row[4]),
            "ageSchedule": schedule,
            "sourceName": "国家免疫规划疫苗儿童免疫程序及说明（2021年版）",
            "sourceFile": source_rel,
            "sourceSheet": ws.title,
            "sourceRow": row_number,
            "sourceYear": 2021,
        })
    wb.close()
    return result


def parse_vaccine_hiv_guidance() -> DatasetResult:
    _, path = vaccine_files()
    result = DatasetResult("VACCINE_HIV_GUIDANCE_2021", "HIV感染母亲所生儿童接种建议", path)
    wb = workbook(path)
    ws = wb.active
    source_rel = str(path.relative_to(ROOT))
    for row_number, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
        vaccine_name = clean(row[0])
        if not vaccine_name:
            continue
        result.records.append({
            "vaccineName": vaccine_name,
            "hivInfectedSymptomatic": clean(row[1]),
            "hivInfectedAsymptomatic": clean(row[2]),
            "hivUnknownSymptomatic": clean(row[3]),
            "hivUnknownAsymptomatic": clean(row[4]),
            "hivUninfected": clean(row[5]),
            "sourceName": "HIV感染母亲所生儿童接种国家免疫规划疫苗建议",
            "sourceFile": source_rel,
            "sourceSheet": ws.title,
            "sourceRow": row_number,
            "sourceYear": 2021,
        })
    wb.close()
    return result


EXCEL_PARSERS: dict[str, Callable[[], DatasetResult]] = {
    "hospitals": parse_hospitals,
    "tertiary": parse_tertiary_hospitals,
    "grades": parse_hospital_grades,
    "drugs": parse_drugs,
    "vaccine-schedule": parse_vaccine_schedule,
    "vaccine-hiv": parse_vaccine_hiv_guidance,
}


def load_dataset(key: str, from_excel: bool = False) -> DatasetResult:
    if not from_excel:
        result = load_processed_dataset(key)
        if result is not None:
            return result
    result = EXCEL_PARSERS[key]()
    if key == "drugs":
        merge_policy_drug_catalog(result)
    return result


def audit_crawl_json() -> list[dict[str, Any]]:
    rows = []
    for path in sorted((ROOT / "data" / "crawl").rglob("*.json")):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception as exc:
            rows.append({"file": str(path.relative_to(ROOT)), "status": "invalid", "error": str(exc)})
            continue
        if not isinstance(payload, dict) or not isinstance(payload.get("items"), list):
            continue
        items = payload["items"]
        seed_urls = sum(
            1 for item in items
            if isinstance(item, dict) and clean(item.get("url")).startswith("seed://")
        )
        rows.append({
            "file": str(path.relative_to(ROOT)),
            "sourceName": payload.get("sourceName", ""),
            "recordCount": payload.get("recordCount"),
            "actualItems": len(items),
            "seedUrlCount": seed_urls,
            "trustedForAutomaticImport": bool(items) and seed_urls == 0,
        })
    return rows


def export_results(results: Iterable[DatasetResult], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    summaries = []
    for result in results:
        records = result.records
        errors = result.errors
        if result.code == "NATIONAL_DRUG_CATALOG_2025":
            policy_file = policy_drug_catalog_file()
            policy_source = str(policy_file.relative_to(ROOT)) if policy_file else ""
            if policy_source:
                policy_export = parse_policy_drug_catalog()
                policy_records = [
                    record for record in result.records
                    if record.get("sourceFile") == policy_source
                ]
                policy_errors = policy_export.errors if policy_export is not None else []
                records = [
                    record for record in result.records
                    if record.get("sourceFile") != policy_source
                ]
                errors = [
                    error for error in result.errors
                    if error.get("sourceFile") != policy_source
                ]
                (output_dir / f"{POLICY_DRUG_CATALOG_CODE}.json").write_text(
                    json.dumps(policy_records, ensure_ascii=False, indent=2), encoding="utf-8"
                )
                (output_dir / f"{POLICY_DRUG_CATALOG_CODE}.errors.json").write_text(
                    json.dumps(policy_errors, ensure_ascii=False, indent=2), encoding="utf-8"
                )
        (output_dir / f"{result.code}.json").write_text(
            json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        (output_dir / f"{result.code}.errors.json").write_text(
            json.dumps(errors, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        summaries.append(result.summary())
    report = {
        "generatedAt": datetime.now().astimezone().isoformat(),
        "mode": "normalized-export-no-database-write",
        "datasets": summaries,
        "crawlSources": audit_crawl_json(),
    }
    (output_dir / "audit-report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def connect_db():
    try:
        import pymysql
    except ImportError as exc:
        raise RuntimeError("执行 --apply 需要安装 pymysql") from exc
    password = os.getenv("HEALTH_DB_PASSWORD")
    if password is None:
        raise RuntimeError("缺少 HEALTH_DB_PASSWORD；禁止在脚本中硬编码数据库密码")
    return pymysql.connect(
        host=os.getenv("HEALTH_DB_HOST", "127.0.0.1"),
        port=int(os.getenv("HEALTH_DB_PORT", "3306")),
        user=os.getenv("HEALTH_DB_USER", "root"),
        password=password,
        database=os.getenv("HEALTH_DB_NAME", "health_portal"),
        charset="utf8mb4",
        autocommit=False,
    )


def assert_migration(conn) -> None:
    with conn.cursor() as cursor:
        cursor.execute(
            "SELECT COUNT(*) FROM information_schema.TABLES "
            "WHERE TABLE_SCHEMA=DATABASE() AND TABLE_NAME='schema_migration'"
        )
        if cursor.fetchone()[0] != 1:
            raise RuntimeError("尚未执行 V002 数据库迁移")
        cursor.execute("SELECT COUNT(*) FROM schema_migration WHERE version='V002'")
        if cursor.fetchone()[0] != 1:
            raise RuntimeError("尚未记录 V002 数据库迁移")


def upsert_many(cursor, sql: str, values: list[tuple[Any, ...]], batch_size: int = 500) -> None:
    for start in range(0, len(values), batch_size):
        cursor.executemany(sql, values[start:start + batch_size])


def apply_results(results: list[DatasetResult]) -> None:
    """显式 --apply 后执行；不删除数据，按来源文件/工作表/行号幂等更新。"""
    conn = connect_db()
    run_id = str(uuid.uuid4())
    try:
        assert_migration(conn)
        with conn.cursor() as cursor:
            for result in results:
                if result.errors:
                    print(json.dumps({
                        "datasetCode": result.code,
                        "warning": "invalid rows were skipped and recorded in error_count",
                        "errorCount": len(result.errors),
                    }, ensure_ascii=False))
                if not result.records:
                    print(json.dumps({
                        "datasetCode": result.code,
                        "warning": "no valid records to import",
                    }, ensure_ascii=False))
                    continue
                if result.code == "PUBLIC_TERTIARY_HOSPITALS":
                    sql = (
                        "INSERT INTO medical_public_tertiary_hospital "
                        "(grade,province,hospital_name,source_name,source_file,source_sheet,source_row) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE "
                        "grade=VALUES(grade),province=VALUES(province),hospital_name=VALUES(hospital_name),"
                        "source_name=VALUES(source_name),updated_at=NOW()"
                    )
                    values = [(
                        r["grade"], r["province"], r["hospitalName"], r["sourceName"],
                        r["sourceFile"], r["sourceSheet"], r["sourceRow"]
                    ) for r in result.records]
                    upsert_many(cursor, sql, values)
                elif result.code == "FUDAN_HOSPITAL_GRADES":
                    sql = (
                        "INSERT INTO medical_hospital_grade "
                        "(grade,hospital_name,source_name,source_file,source_sheet,source_row) "
                        "VALUES (%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE "
                        "grade=VALUES(grade),hospital_name=VALUES(hospital_name),source_name=VALUES(source_name)"
                    )
                    values = [(
                        r["grade"], r["hospitalName"], r["sourceName"], r["sourceFile"],
                        r["sourceSheet"], r["sourceRow"]
                    ) for r in result.records]
                    upsert_many(cursor, sql, values)
                elif result.code == "NATIONAL_DRUG_CATALOG_2025":
                    sql = (
                        "INSERT INTO medical_drug_catalog "
                        "(category_code,category_name,drug_number,drug_name,dosage_form,insurance_type,remark,"
                        "catalog_year,source_file,source_sheet,source_row) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) "
                        "ON DUPLICATE KEY UPDATE category_code=VALUES(category_code),category_name=VALUES(category_name),"
                        "drug_number=VALUES(drug_number),drug_name=VALUES(drug_name),dosage_form=VALUES(dosage_form),"
                        "insurance_type=VALUES(insurance_type),remark=VALUES(remark),updated_at=NOW()"
                    )
                    values = [(
                        r["categoryCode"], r["categoryName"], r["drugNumber"], r["drugName"],
                        r["dosageForm"] or None, r["insuranceType"] or None, r["remark"] or None,
                        r["catalogYear"], r["sourceFile"], r["sourceSheet"], r["sourceRow"]
                    ) for r in result.records]
                    upsert_many(cursor, sql, values)
                elif result.code == "VACCINE_SCHEDULE_2021":
                    sql = (
                        "INSERT INTO medical_vaccine_schedule "
                        "(preventable_disease,vaccine_name,route,dose,abbreviation,age_schedule_json,source_name,"
                        "source_file,source_sheet,source_row,source_year) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) "
                        "ON DUPLICATE KEY UPDATE preventable_disease=VALUES(preventable_disease),"
                        "vaccine_name=VALUES(vaccine_name),route=VALUES(route),dose=VALUES(dose),"
                        "abbreviation=VALUES(abbreviation),age_schedule_json=VALUES(age_schedule_json)"
                    )
                    values = [(
                        r["preventableDisease"], r["vaccineName"], r["route"], r["dose"], r["abbreviation"],
                        json.dumps(r["ageSchedule"], ensure_ascii=False), r["sourceName"], r["sourceFile"],
                        r["sourceSheet"], r["sourceRow"], r["sourceYear"]
                    ) for r in result.records]
                    upsert_many(cursor, sql, values)
                elif result.code == "VACCINE_HIV_GUIDANCE_2021":
                    sql = (
                        "INSERT INTO medical_vaccine_hiv_guidance "
                        "(vaccine_name,hiv_infected_symptomatic,hiv_infected_asymptomatic,hiv_unknown_symptomatic,"
                        "hiv_unknown_asymptomatic,hiv_uninfected,source_name,source_file,source_sheet,source_row,source_year) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON DUPLICATE KEY UPDATE "
                        "vaccine_name=VALUES(vaccine_name),hiv_infected_symptomatic=VALUES(hiv_infected_symptomatic),"
                        "hiv_infected_asymptomatic=VALUES(hiv_infected_asymptomatic),"
                        "hiv_unknown_symptomatic=VALUES(hiv_unknown_symptomatic),"
                        "hiv_unknown_asymptomatic=VALUES(hiv_unknown_asymptomatic),hiv_uninfected=VALUES(hiv_uninfected)"
                    )
                    values = [(
                        r["vaccineName"], r["hivInfectedSymptomatic"], r["hivInfectedAsymptomatic"],
                        r["hivUnknownSymptomatic"], r["hivUnknownAsymptomatic"], r["hivUninfected"],
                        r["sourceName"], r["sourceFile"], r["sourceSheet"], r["sourceRow"], r["sourceYear"]
                    ) for r in result.records]
                    upsert_many(cursor, sql, values)
                elif result.code == "HOSPITAL_DIRECTORY_2024":
                    sql = (
                        "INSERT INTO medical_hospital "
                        "(source_record_no,name,alias_name,province,city,district,address,level,type,founded_year,"
                        "operation_mode,is_insurance,bed_count,annual_visits,medical_staff_count,departments,phone,"
                        "email,postal_code,introduction,source_name,source_file,source_sheet,source_row) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) "
                        "ON DUPLICATE KEY UPDATE name=VALUES(name),alias_name=VALUES(alias_name),province=VALUES(province),"
                        "city=VALUES(city),district=VALUES(district),address=VALUES(address),level=VALUES(level),"
                        "type=VALUES(type),phone=VALUES(phone),updated_at=NOW()"
                    )
                    values = [(
                        r["sourceRecordNo"], r["name"], r["aliasName"], r["province"], r["city"], r["district"],
                        r["address"], r["level"], r["type"], r["foundedYear"], r["operationMode"], r["isInsurance"],
                        r["bedCount"], r["annualVisits"], r["medicalStaffCount"], r["departments"], r["phone"],
                        r["email"], r["postalCode"], r["introduction"], r["sourceName"], r["sourceFile"],
                        r["sourceSheet"], r["sourceRow"]
                    ) for r in result.records]
                    upsert_many(cursor, sql, values)
                else:
                    raise RuntimeError(f"没有实现 {result.code} 的数据库映射")

                cursor.execute(
                    "INSERT INTO data_resource_dataset "
                    "(dataset_code,dataset_name,dataset_type,source_name,source_file,record_count,duplicate_count,"
                    "error_count,update_status,last_imported_at) VALUES (%s,%s,'DATABASE',%s,%s,%s,%s,%s,'SUCCESS',NOW()) "
                    "ON DUPLICATE KEY UPDATE dataset_name=VALUES(dataset_name),source_name=VALUES(source_name),"
                    "source_file=VALUES(source_file),record_count=VALUES(record_count),"
                    "duplicate_count=VALUES(duplicate_count),error_count=VALUES(error_count),"
                    "update_status='SUCCESS',last_imported_at=NOW()",
                    (result.code, result.name, result.name, str(result.source_file.relative_to(ROOT)),
                     len(result.records), result.duplicate_count, len(result.errors)),
                )
        conn.commit()
        print(json.dumps({"status": "success", "runId": run_id}, ensure_ascii=False))
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--dataset", action="append", choices=sorted(EXCEL_PARSERS),
        help="只处理指定数据集；可重复使用。默认处理全部。",
    )
    parser.add_argument("--from-excel", action="store_true", help="从 data/external-import 的 Excel 重新解析；默认直接读取标准化 JSON")
    parser.add_argument("--export", action="store_true", help="输出标准化 JSON 和错误报告")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="标准化输出目录")
    parser.add_argument("--apply", action="store_true", help="显式写入数据库；默认绝不写库")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    selected = args.dataset or list(EXCEL_PARSERS)
    results = []
    for key in selected:
        result = load_dataset(key, args.from_excel)
        results.append(result)
        print(json.dumps(result.summary(), ensure_ascii=False))

    crawl_audit = audit_crawl_json()
    print(json.dumps({"crawlSources": crawl_audit}, ensure_ascii=False))

    if args.export:
        export_results(results, args.output)
        print(json.dumps({"exportedTo": str(args.output)}, ensure_ascii=False))
    if args.apply:
        apply_results(results)
    else:
        print(json.dumps({"databaseWrite": False, "mode": "dry-run"}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(json.dumps({"status": "error", "message": str(exc)}, ensure_ascii=False), file=sys.stderr)
        raise SystemExit(1)

